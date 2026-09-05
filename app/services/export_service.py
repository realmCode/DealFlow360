"""Report export — PDF A7.2 requires "Export options: PDF / XLS".

These are the only binary responses in the API; everything else is JSON.

`openpyxl` and `reportlab` are imported lazily inside their branches so a
deployment that never exports does not pay the import cost, and — more usefully
— so a missing optional dependency produces a clear 400 naming the package
rather than breaking application start-up.
"""

from __future__ import annotations

import csv
import io
from datetime import UTC, datetime
from typing import Any, Sequence

from app.errors import BusinessRuleError

CSV_MEDIA_TYPE = "text/csv; charset=utf-8"
XLSX_MEDIA_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
PDF_MEDIA_TYPE = "application/pdf"

SUPPORTED_FORMATS = ("csv", "xlsx", "pdf")


def _humanise(header: str) -> str:
    return header.replace("_", " ").title()


def _stringify(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


class ExportService:
    @staticmethod
    def filename(report_name: str, fmt: str) -> str:
        stamp = datetime.now(UTC).strftime("%Y%m%d-%H%M%S")
        return f"dealflow360-{report_name}-{stamp}.{fmt}"

    @staticmethod
    def media_type(fmt: str) -> str:
        return {
            "csv": CSV_MEDIA_TYPE,
            "xlsx": XLSX_MEDIA_TYPE,
            "pdf": PDF_MEDIA_TYPE,
        }[fmt]

    # ------------------------------------------------------------------ csv
    @staticmethod
    def to_csv(headers: Sequence[str], rows: list[dict[str, Any]]) -> bytes:
        buffer = io.StringIO(newline="")
        writer = csv.writer(buffer, quoting=csv.QUOTE_MINIMAL)
        writer.writerow([_humanise(h) for h in headers])
        for row in rows:
            writer.writerow([_stringify(row.get(h)) for h in headers])
        # utf-8-sig so Excel opens the file with the right encoding on
        # a double-click instead of mangling non-ASCII customer names.
        return buffer.getvalue().encode("utf-8-sig")

    # ----------------------------------------------------------------- xlsx
    @staticmethod
    def to_xlsx(
        headers: Sequence[str],
        rows: list[dict[str, Any]],
        *,
        title: str,
        meta: dict[str, Any] | None = None,
    ) -> bytes:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ModuleNotFoundError as exc:  # pragma: no cover - env dependent
            raise BusinessRuleError(
                "XLSX export requires the 'openpyxl' package.",
                code="EXPORT_DEPENDENCY_MISSING",
                details={"package": "openpyxl", "format": "xlsx"},
            ) from exc

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = title[:31] or "Report"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", start_color="1F3864")

        sheet.append([_humanise(h) for h in headers])
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in rows:
            sheet.append([_stringify(row.get(h)) for h in headers])

        # Width from the widest cell, so a reviewer does not have to resize
        # every column before the numbers are legible.
        for index, header in enumerate(headers, start=1):
            longest = max(
                [len(_humanise(header))]
                + [len(_stringify(row.get(header))) for row in rows]
                or [10]
            )
            sheet.column_dimensions[get_column_letter(index)].width = min(
                48, longest + 2
            )
        sheet.freeze_panes = "A2"

        if meta:
            info = workbook.create_sheet("Filters")
            info.append(["Filter", "Value"])
            for cell in info[1]:
                cell.font = header_font
                cell.fill = header_fill
            for key, value in meta.items():
                info.append([_humanise(str(key)), _stringify(value)])
            info.column_dimensions["A"].width = 28
            info.column_dimensions["B"].width = 48

        stream = io.BytesIO()
        workbook.save(stream)
        return stream.getvalue()

    # ------------------------------------------------------------------ pdf
    @staticmethod
    def to_pdf(
        headers: Sequence[str],
        rows: list[dict[str, Any]],
        *,
        title: str,
        meta: dict[str, Any] | None = None,
    ) -> bytes:
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib.units import mm
            from reportlab.platypus import (
                Paragraph,
                SimpleDocTemplate,
                Spacer,
                Table,
                TableStyle,
            )
        except ModuleNotFoundError as exc:  # pragma: no cover - env dependent
            raise BusinessRuleError(
                "PDF export requires the 'reportlab' package.",
                code="EXPORT_DEPENDENCY_MISSING",
                details={"package": "reportlab", "format": "pdf"},
            ) from exc

        stream = io.BytesIO()
        doc = SimpleDocTemplate(
            stream,
            pagesize=landscape(A4),
            leftMargin=12 * mm,
            rightMargin=12 * mm,
            topMargin=12 * mm,
            bottomMargin=12 * mm,
            title=title,
            author="DealFlow360",
        )
        styles = getSampleStyleSheet()
        story: list[Any] = [
            Paragraph(f"DealFlow360 — {title}", styles["Title"]),
            Paragraph(
                f"Generated {datetime.now(UTC).strftime('%Y-%m-%d %H:%M UTC')}",
                styles["Normal"],
            ),
            Spacer(1, 6 * mm),
        ]

        if meta:
            applied = ", ".join(
                f"{_humanise(str(k))}: {_stringify(v)}"
                for k, v in meta.items()
                if v not in (None, "")
            )
            if applied:
                story.append(Paragraph(f"<b>Filters</b> — {applied}", styles["Normal"]))
                story.append(Spacer(1, 4 * mm))

        if not rows:
            story.append(
                Paragraph(
                    "No data matched the selected filters.", styles["Italic"]
                )
            )
        else:
            data = [[_humanise(h) for h in headers]]
            data += [[_stringify(row.get(h)) for h in headers] for row in rows]
            table = Table(data, repeatRows=1, hAlign="LEFT")
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F3864")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, -1), 7),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#BFBFBF")),
                        (
                            "ROWBACKGROUNDS",
                            (0, 1),
                            (-1, -1),
                            [colors.white, colors.HexColor("#F2F2F2")],
                        ),
                        ("TOPPADDING", (0, 0), (-1, -1), 3),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                    ]
                )
            )
            story.append(table)

        doc.build(story)
        return stream.getvalue()

    # ---------------------------------------------------------------- entry
    @classmethod
    def render(
        cls,
        *,
        fmt: str,
        report_name: str,
        headers: Sequence[str],
        rows: list[dict[str, Any]],
        meta: dict[str, Any] | None = None,
    ) -> tuple[bytes, str, str]:
        """Return ``(body, media_type, filename)``."""
        if fmt not in SUPPORTED_FORMATS:
            raise BusinessRuleError(
                f"Unsupported export format {fmt!r}.",
                code="UNSUPPORTED_EXPORT_FORMAT",
                details={"format": fmt, "supported": list(SUPPORTED_FORMATS)},
            )

        title = _humanise(report_name)
        if fmt == "csv":
            body = cls.to_csv(headers, rows)
        elif fmt == "xlsx":
            body = cls.to_xlsx(headers, rows, title=title, meta=meta)
        else:
            body = cls.to_pdf(headers, rows, title=title, meta=meta)

        return body, cls.media_type(fmt), cls.filename(report_name, fmt)
